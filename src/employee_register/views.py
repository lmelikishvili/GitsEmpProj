from django.shortcuts import redirect, render
from .forms import EmployeeForm
from .models import Employee
from .filters import OrderFilter
from django.http import HttpResponse
import datetime
import csv, codecs
import xlwt
from openpyxl import Workbook


# Create your views here.

def employee_list(request):
    emp = Employee.objects.all()
    myFilter = OrderFilter(request.GET, queryset=Employee.objects.all())
    emp = myFilter.qs
    context = {'employee_list': emp, 'myFilter': myFilter}
    return render(request, 'employee_register/employee_list.html', context)
    
def vacation_view(request):
    return render(request, 'employee_register/vacation/vacation.html')

def employee_form(request, id=0):
    if request.method == "GET":
        if id == 0:
            form = EmployeeForm()
        else:
            employee = Employee.objects.get(pk=id)
            form = EmployeeForm(instance = employee)
        return render(request, 'employee_register/employee_form.html', {'form': form})
    else:
        if id == 0:
            form = EmployeeForm(request.POST)
        else:
            employee = Employee.objects.get(pk=id)
            form = EmployeeForm(request.POST, instance = employee)
        if form.is_valid():
            form.save()
        return redirect('employee_list')

def employee_delete(request, id):
    employee = Employee.objects.get(pk=id)
    employee.delete()
    return redirect('employee_list')

def export_csv(request):
    response = HttpResponse(content_type = 'text/csv')
    writer = csv.writer(response)
    writer.writerow(['სახელი, გვარი', 'მობილური', 'მეილი', 'თანამდებობა'])
    response.write(codecs.BOM_UTF8)
    for emp in Employee.objects.all():
        writer.writerow([emp.fullname, emp.mobile, emp.email, emp.position])
    
    response['Content-Disposition'] = 'attachment; filename="empys.csv"'
    return response


def export_excell(request):
    emp_set = Employee.objects.all()
    response = HttpResponse(content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="empys.xlsx"'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Empys'
    
    columns = ['სახელი, გვარი', 'მობილური', 'მეილი', 'თანამდებობა']
    row_num = 1
    for col_num, column_title in enumerate(columns, row_num):
        cell = worksheet.cell(row=row_num, column = col_num)
        cell.value = column_title
    
    for e in emp_set:
        row_num += 1
        row = [
            e.fullname,
            e.mobile,
            e.email,
            e.position.title,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row = row_num, column = col_num)
            cell.value = cell_value

    workbook.save(response)
    return response
